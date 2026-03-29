from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook

from config.settings import DEFAULT_EXPORT_PATH, ensure_directories
from repositories.lineas_repository import LineasRepository
from repositories.ordenes_repository import OrdenesRepository


class ExportacionService:
    """
    Servicio para exportar resultados del MVP a Excel.
    """

    @staticmethod
    def _crear_hoja_secuencia(wb: Workbook, resultado_secuencia: dict[str, Any]) -> None:
        ws = wb.active
        ws.title = "Secuencia"

        headers = [
            "ID Línea",
            "Código Línea",
            "Nombre Línea",
            "Posición",
            "Tipo Registro",
            "ID OP",
            "Número OP",
            "ID Limpieza",
            "Categoría Origen",
            "Categoría Destino",
            "Acción",
            "Justificación",
            "Editado Manual",
        ]
        ws.append(headers)

        for registro in resultado_secuencia.get("registros", []):
            numero_op = ""
            if registro.get("id_op"):
                op = OrdenesRepository.get_by_id(registro["id_op"])
                numero_op = op.numero_op if op else ""

            ws.append(
                [
                    registro.get("id_linea"),
                    registro.get("codigo_linea"),
                    registro.get("nombre_linea"),
                    registro.get("posicion"),
                    registro.get("tipo_registro"),
                    registro.get("id_op"),
                    numero_op,
                    registro.get("id_limpieza"),
                    registro.get("categoria_origen"),
                    registro.get("categoria_destino"),
                    registro.get("accion_transicion"),
                    registro.get("justificacion"),
                    registro.get("editado_manual"),
                ]
            )

    @staticmethod
    def _crear_hoja_resumen(wb: Workbook, resultado_secuencia: dict[str, Any]) -> None:
        ws = wb.create_sheet(title="Resumen")

        headers = [
            "Código Línea",
            "Nombre Línea",
            "Producciones",
            "Limpiezas",
            "Flush",
            "Total Registros",
        ]
        ws.append(headers)

        resumen: dict[str, dict[str, Any]] = {}

        for linea in LineasRepository.list_all(include_inactive=False):
            resumen[linea.codigo_linea] = {
                "nombre_linea": linea.nombre_linea,
                "producciones": 0,
                "limpiezas": 0,
                "flush": 0,
                "total": 0,
            }

        for registro in resultado_secuencia.get("registros", []):
            codigo = registro["codigo_linea"]
            if codigo not in resumen:
                continue

            resumen[codigo]["total"] += 1

            if registro["tipo_registro"] == "PRODUCCION":
                resumen[codigo]["producciones"] += 1
            elif registro["tipo_registro"] == "LIMPIEZA":
                resumen[codigo]["limpiezas"] += 1
            elif registro["tipo_registro"] == "FLUSH":
                resumen[codigo]["flush"] += 1

        for codigo_linea, data in resumen.items():
            ws.append(
                [
                    codigo_linea,
                    data["nombre_linea"],
                    data["producciones"],
                    data["limpiezas"],
                    data["flush"],
                    data["total"],
                ]
            )

    @staticmethod
    def _crear_hoja_alertas(wb: Workbook, alertas: dict[str, list[str]]) -> None:
        ws = wb.create_sheet(title="Alertas")
        ws.append(["Tipo", "Mensaje"])

        for error in alertas.get("errores", []):
            ws.append(["ERROR", error])

        for alerta in alertas.get("alertas", []):
            ws.append(["ALERTA", alerta])

    @staticmethod
    def exportar_resultado(
        resultado_secuencia: dict[str, Any],
        alertas: dict[str, list[str]],
        output_path: str | Path | None = None,
    ) -> Path:
        ensure_directories()

        target_path = Path(output_path) if output_path else DEFAULT_EXPORT_PATH

        wb = Workbook()
        ExportacionService._crear_hoja_secuencia(wb, resultado_secuencia)
        ExportacionService._crear_hoja_resumen(wb, resultado_secuencia)
        ExportacionService._crear_hoja_alertas(wb, alertas)

        wb.save(target_path)
        return target_path
